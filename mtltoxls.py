import re
import glob
import sys
import os
from openpyxl import Workbook
from pprint import pprint

re_key = re.compile(r'(\S*)\s*=\s*\{\s*Name\s*=\s*(\S*)\s*\n')
re_parameter = re.compile(r'(\{(?:\s*.* = .*\s*\n)*\})')
re_key_value_parameter = re.compile(r'\s*(?P<key>\S*) = (?P<value>\S*)( .*)?\s*\n')


def parse_mtl(fname):
    with open(fname, 'r', encoding='utf8') as fd:
        mtl = fd.read()

    # Find all parameter sections in curly brackets like for example:
    #    {
    #        Name = PTC_INITIAL_BEND_Y_FACTOR
    #        Type = Real
    #        Default = 5.000000e-01
    #        Access = Full
    #    },
    m = re.findall(re_parameter, mtl)

    # For each parameter set, extract the list of Key-Value pairs like "Name = PTC_INITIAL..."
    parameters = []
    for parameter in m:
        matches = re.findall(re_key_value_parameter, parameter)
        if matches:
            parameters.append(matches)

    # Extract all parameter sets into a list of dictionaries
    # Adding a new dict key: 'Unit' which comes from the 'Default' value field.
    parameter_dicts = []
    for parameter in parameters:
        pd = {'Unit': None}
        for key, value, unit in parameter:
            pd.update({key: value})
            if unit is not '':
                pd.update({'Unit': unit.strip()})
        parameter_dicts.append(pd)

    # Convert the string values of all of the parameters 'Default' field into a proper python
    # data type - so that it can later be stored appropriately in the spreadsheet
    for pd in parameter_dicts:
        value = pd['Default']
        pdtype = pd['Type']
        if pdtype == 'Real':
            value = float(value)
        elif pdtype == 'Integer':
            value = int(value)
        elif pdtype == 'String':
            # TODO: hack alert - if only I knew better regexp I could avoid this...
            if pd['Unit'] is not None:
                value = "{} {}".format(str(value), str(pd['Unit'])).strip()
                pd['Unit'] = ''
            else:
                value = str(value).strip()
            value = value.strip('\'').strip('\"')
        pd['Default'] = value

    # make a dictionary of parameters where the parameter Name is the key
    material_params = {}
    for pd in parameter_dicts:
        material_params.update({pd['Name']: pd})

    #pprint(material_params)

    # Scan file content again to extract the material ID and Name.
    material_key = re.findall(re_key, mtl[1:])
    if material_key:
        key, material = material_key[0]
    else:
        key = 'unknown_key'
        material = 'unknown_material'

    # Generate final top-level dictionary view of the entire file including material Name and ID.
    result = {key: {material: material_params}}
    return result


class ParameterRowIndex:
    """Maintain a dictionary of parameter name -> parameter row index"""
    def __init__(self, offset=1):
        self._parameter_index = {}
        self._current_offset = offset

    def add_parameters(self, parameters: list):
        """Add new parameters to the index data. Only new parameters from the 'parameters' list are added with
        an incremental index
        :parameters is a list of dictionaries with fields for Type and Access"""
        param_type_access = [(p, parameters[p]['Type'], parameters[p]['Access']) for p in parameters.keys()]
        for param, param_type, param_access in param_type_access:
            if param not in self._parameter_index:
                self._parameter_index.update({param: {'Offset': self._current_offset, 'Type': param_type, 'Access': param_access}})
                self._current_offset += 1

    @property
    def parameter_index(self):
        return self._parameter_index


class StoreSpreadsheet:
    def __init__(self, fname):
        self._fname = fname
        self._workbook = Workbook()
        self._param_row = ParameterRowIndex(offset=3)
        self._current_col = 4
        self._workbook.active.cell(1,1, "Material ID:")
        self._workbook.active.cell(2,1, "Material Name:")

    def store_material_parameters(self, materials):
        for material_id in materials.keys():
            # Work sheet
            ws = self._workbook.active
            ws.cell(1, self._current_col, material_id)
            ws.merge_cells(start_row=1, end_row=1, start_column=self._current_col, end_column=self._current_col + 1)

            for material in materials[material_id].keys():

                # Header: material name
                ws.cell(2, self._current_col, material)
                ws.merge_cells(start_row=2, end_row=2, start_column=self._current_col, end_column=self._current_col+1)

                parameters = materials[material_id][material]
                self._param_row.add_parameters(parameters)
                for param in parameters:
                    ws.cell(self._param_row.parameter_index[param]['Offset'], self._current_col, parameters[param]['Default'])
                    ws.cell(self._param_row.parameter_index[param]['Offset'], self._current_col + 1, parameters[param]['Unit'])
                self._current_col += 2

    def update_parameter_column(self):
        wb = self._workbook.active
        for p in self._param_row.parameter_index:
            row = self._param_row.parameter_index[p]['Offset']
            wb.cell(row, 1, p)
            wb.cell(row, 2, self._param_row.parameter_index[p]['Type'])
            wb.cell(row, 3, self._param_row.parameter_index[p]['Access'])

    def save(self):
        self.update_parameter_column()
        self._workbook.save(self._fname)


def main():
    path = sys.argv[1]
    full_path = os.path.abspath(path)
    print(F"Searching dir for mtl files: \"{full_path}\"")
    fnames = glob.glob(F"{full_path}/*.mtl")
    pprint(fnames)

    materials = []
    for fname in fnames:
        materials.append(parse_mtl(fname))

    spreadsheet_fname = str(sys.argv[2])
    print(F"Creating spreadsheet: {spreadsheet_fname}")
    s = StoreSpreadsheet(spreadsheet_fname)
    for material in materials:
        print(F"storing material: \"{material}\" in spreadsheet")
        s.store_material_parameters(material)
    s.save()


if __name__=="__main__":
    main()
